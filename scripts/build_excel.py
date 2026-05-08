"""Build a formatted Excel workbook from ingredients_raw.json.

Sheets:
  - "Par couleur"     : 4 columns (Vert, Jaune, Orange, Rouge), alphabetical
  - "Tous"            : full list with color, name, traduction (alphabetical)
  - "Vert" / "Jaune" / "Orange" / "Rouge" : per-color detail sheet

Cells are color-coded; a legend is rendered at the top of the main sheet.
"""
import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

ROOT = Path(__file__).parent
RAW = ROOT / "ingredients_raw.json"
OUT_XLSX = ROOT / "incibeauty_ingredients.xlsx"

COLOR_ORDER = ["Vert", "Jaune", "Orange", "Rouge"]
FILLS = {
    "Vert":   PatternFill("solid", fgColor="C6EFCE"),
    "Jaune":  PatternFill("solid", fgColor="FFEB9C"),
    "Orange": PatternFill("solid", fgColor="FFD9A8"),
    "Rouge":  PatternFill("solid", fgColor="FFC7CE"),
}
HEADER_FILLS = {
    "Vert":   PatternFill("solid", fgColor="2E7D32"),
    "Jaune":  PatternFill("solid", fgColor="F9A825"),
    "Orange": PatternFill("solid", fgColor="EF6C00"),
    "Rouge":  PatternFill("solid", fgColor="C62828"),
}
HEADER_FONT = Font(bold=True, color="FFFFFF", size=12)
TITLE_FONT = Font(bold=True, size=16, color="1F2937")
SUBTITLE_FONT = Font(italic=True, size=10, color="555555")
LINK_FONT = Font(color="0563C1", underline="single")
THIN = Side(style="thin", color="D0D0D0")
BORDER = Border(top=THIN, left=THIN, right=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def load() -> list[dict]:
    return json.loads(RAW.read_text(encoding="utf-8"))


def by_color(rows: list[dict]) -> dict[str, list[dict]]:
    out: dict[str, list[dict]] = {c: [] for c in COLOR_ORDER}
    for r in rows:
        out.setdefault(r["color"], []).append(r)
    for c in out:
        out[c].sort(key=lambda r: r["name"].upper())
    return out


def write_legend_and_title(ws, total: int, counts: dict[str, int]) -> int:
    ws["A1"] = "INCI Beauty - Ingrédients cosmétiques par classification couleur"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLOR_ORDER))
    ws.row_dimensions[1].height = 28

    sub = (
        f"Source: incibeauty.com/ingredients - {total} ingrédients uniques "
        f"(Vert: {counts.get('Vert',0)}, Jaune: {counts.get('Jaune',0)}, "
        f"Orange: {counts.get('Orange',0)}, Rouge: {counts.get('Rouge',0)})"
    )
    ws["A2"] = sub
    ws["A2"].font = SUBTITLE_FONT
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(COLOR_ORDER))
    ws.row_dimensions[2].height = 18

    legend_row = 3
    legend_labels = {
        "Vert": "Vert - Excellent",
        "Jaune": "Jaune - Satisfaisant",
        "Orange": "Orange - Médiocre",
        "Rouge": "Rouge - À éviter",
    }
    for i, c in enumerate(COLOR_ORDER):
        cell = ws.cell(row=legend_row, column=i + 1, value=legend_labels[c])
        cell.fill = FILLS[c]
        cell.font = Font(bold=True, color="222222")
        cell.alignment = CENTER
        cell.border = BORDER
    ws.row_dimensions[legend_row].height = 22
    return legend_row + 2  # leave a blank row


def build_par_couleur(ws, grouped: dict[str, list[dict]], counts: dict[str, int]) -> None:
    total = sum(counts.values())
    start = write_legend_and_title(ws, total, counts)

    # header row
    for i, c in enumerate(COLOR_ORDER):
        cell = ws.cell(row=start, column=i + 1, value=f"{c} ({counts.get(c, 0)})")
        cell.fill = HEADER_FILLS[c]
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER
    ws.row_dimensions[start].height = 24

    max_len = max(len(grouped[c]) for c in COLOR_ORDER)
    for col_idx, c in enumerate(COLOR_ORDER, start=1):
        col_rows = grouped[c]
        for r_idx in range(max_len):
            cell = ws.cell(row=start + 1 + r_idx, column=col_idx)
            if r_idx < len(col_rows):
                item = col_rows[r_idx]
                cell.value = item["name"]
                cell.fill = FILLS[c]
                if item.get("url"):
                    cell.hyperlink = item["url"]
                    cell.font = LINK_FONT
            cell.alignment = LEFT
            cell.border = BORDER

    # column widths
    for i, c in enumerate(COLOR_ORDER):
        col_letter = get_column_letter(i + 1)
        names = [r["name"] for r in grouped[c]]
        max_w = max([len(n) for n in names] + [len(c) + 8, 18])
        ws.column_dimensions[col_letter].width = min(max_w + 2, 60)

    ws.freeze_panes = ws.cell(row=start + 1, column=1)
    ws.sheet_view.showGridLines = False


def build_per_color_sheet(wb: Workbook, color: str, rows: list[dict]) -> None:
    ws = wb.create_sheet(color)
    headers = ["#", "Ingrédient (INCI)", "Traduction FR"]
    title = ws.cell(row=1, column=1, value=f"Ingrédients - {color} ({len(rows)})")
    title.font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws.row_dimensions[1].height = 24

    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=i, value=h)
        cell.fill = HEADER_FILLS[color]
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER
    ws.row_dimensions[2].height = 22

    for r_idx, item in enumerate(rows, start=3):
        idx_cell = ws.cell(row=r_idx, column=1, value=r_idx - 2)
        name_cell = ws.cell(row=r_idx, column=2, value=item["name"])
        tr_cell = ws.cell(row=r_idx, column=3, value=item.get("translation", ""))
        for c in (idx_cell, name_cell, tr_cell):
            c.border = BORDER
            c.fill = FILLS[color]
        idx_cell.alignment = CENTER
        name_cell.alignment = LEFT
        tr_cell.alignment = LEFT
        if item.get("url"):
            name_cell.hyperlink = item["url"]
            name_cell.font = LINK_FONT

    widths = [6, 60, 50]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A3"
    ws.sheet_view.showGridLines = False
    ws.auto_filter.ref = f"A2:C{len(rows) + 2}"


def build_tous(wb: Workbook, rows: list[dict]) -> None:
    ws = wb.create_sheet("Tous")
    rows_sorted = sorted(rows, key=lambda r: r["name"].upper())

    title = ws.cell(row=1, column=1, value=f"Tous les ingrédients ({len(rows_sorted)})")
    title.font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
    ws.row_dimensions[1].height = 24

    headers = ["#", "Couleur", "Ingrédient (INCI)", "Traduction FR"]
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=i, value=h)
        cell.fill = PatternFill("solid", fgColor="263238")
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER
    ws.row_dimensions[2].height = 22

    for r_idx, item in enumerate(rows_sorted, start=3):
        ws.cell(row=r_idx, column=1, value=r_idx - 2).alignment = CENTER
        c_color = ws.cell(row=r_idx, column=2, value=item["color"])
        c_name = ws.cell(row=r_idx, column=3, value=item["name"])
        c_tr = ws.cell(row=r_idx, column=4, value=item.get("translation", ""))
        fill = FILLS[item["color"]]
        for cell in (ws.cell(row=r_idx, column=1), c_color, c_name, c_tr):
            cell.fill = fill
            cell.border = BORDER
        c_color.alignment = CENTER
        c_name.alignment = LEFT
        c_tr.alignment = LEFT
        if item.get("url"):
            c_name.hyperlink = item["url"]
            c_name.font = LINK_FONT

    widths = [6, 12, 60, 50]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A3"
    ws.sheet_view.showGridLines = False
    ws.auto_filter.ref = f"A2:D{len(rows_sorted) + 2}"


def main() -> None:
    rows = load()
    grouped = by_color(rows)
    counts = {c: len(grouped[c]) for c in COLOR_ORDER}

    wb = Workbook()
    main_ws = wb.active
    main_ws.title = "Par couleur"

    build_par_couleur(main_ws, grouped, counts)
    build_tous(wb, rows)
    for c in COLOR_ORDER:
        build_per_color_sheet(wb, c, grouped[c])

    wb.save(OUT_XLSX)
    print(f"Wrote {OUT_XLSX} ({OUT_XLSX.stat().st_size/1024:.0f} KiB)")
    print(f"Counts: {counts}, total={sum(counts.values())}")


if __name__ == "__main__":
    main()
